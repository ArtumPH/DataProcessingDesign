# excel_handler.py
from openpyxl import load_workbook
from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView

def load_data_from_excel(file_name, table_widget):
    workbook = load_workbook(file_name, data_only=True)
    sheet = workbook.active

    table_widget.setRowCount(sheet.max_row)
    table_widget.setColumnCount(sheet.max_column)

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_index, value in enumerate(row):
            item = QTableWidgetItem(str(value))
            table_widget.setItem(row_index - 1, col_index, item)

    # Resize columns to fit contents
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.ResizeToContents)
