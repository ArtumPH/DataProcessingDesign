# TableWidget.py
from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView, QVBoxLayout, QWidget, QTextEdit, QMessageBox
from openpyxl import load_workbook
import pandas as pd

class ExcelTableWidget(QTableWidget):
    def __init__(self):
        super().__init__()

        # Add selected_data_display attribute
        #self.selected_data_display = QTextEdit(self)
        self.display_text = 'test contents'

    def load_data_from_excel(self, file_name):
        workbook = load_workbook(file_name, data_only=True)
        sheet = workbook.active

        self.setRowCount(sheet.max_row)
        self.setColumnCount(sheet.max_column)

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_index, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.setItem(row_index - 1, col_index, item)

        # Resize columns to fit contents
        header = self.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    #Unfinished function
    def get_selected_data_table_form(self):
        selected_items = self.selectedItems()

        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select some cells before reading data.')
            return

        #rows = set()
        #cols = set()
        rows = list()
        cols = list()

        for item in selected_items:
            rows.add(item.row())
            cols.add(item.column())

        #selected_rows = sorted(list(rows))
        #selected_cols = sorted(list(cols))
        selected_rows = rows
        selected_cols = cols

        selected_data_matrix = []
        for row in selected_rows:
            row_data = [self.item(row, col) for col in selected_cols]
            selected_data_matrix.append(row_data)

        selected_data_columns = [f'{col+1}' for col in selected_cols]

        return selected_data_matrix, selected_data_columns

    def get_selected_data(self):
        selected_items = self.selectedItems()

        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select some cells before reading data.')
            return

        #rows = list()
        #cols = list()
        #Is it right and safe to use set rather than list? Set dont allowed repeated elements.
        rows = set()
        cols = set()

        for item in selected_items:
            rows.add(item.row())
            cols.add(item.column())

        #selected_rows = rows
        #selected_cols = cols
        selected_rows = sorted(list(rows))
        selected_cols = sorted(list(cols))

        selected_data_matrix = []
        for row in selected_rows:
            row_data = [self.item(row, col).text() for col in selected_cols]
            selected_data_matrix.append(row_data)

        selected_data_columns = [f'Col{col+1}' for col in selected_cols]

        return selected_data_matrix, selected_data_columns

    def display_selected_data(self, data_matrix, columns):
        self.display_text = f'Selected Data in Matrix Form:\n\n{pd.DataFrame(data_matrix, columns=columns)}'
        #self.selected_data_display.setPlainText(display_text)
        return self.display_text

class ExcelTextEdit(QTextEdit):
    def __init__(self):
        super().__init__()
