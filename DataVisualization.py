import sys
from PyQt6.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QPushButton, QFileDialog, QMessageBox, QListWidget, QComboBox, QHBoxLayout, QTextEdit, QHeaderView

from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        # Entire layout using QHBoxLayout
        self.layout = QHBoxLayout(self.central_widget)

        # Left side layout
        self.layout_left = QVBoxLayout()

        self.table_widget = QTableWidget(self)
        self.layout_left.addWidget(self.table_widget)

        self.load_button = QPushButton('Load Excel', self)
        self.load_button.clicked.connect(self.load_excel)
        self.layout_left.addWidget(self.load_button)

        self.read_button = QPushButton('Read Selected Data', self)
        self.read_button.clicked.connect(self.read_selected_data)
        self.layout_left.addWidget(self.read_button)

        # Add stretch factor between left and right sides
        self.layout.addLayout(self.layout_left, stretch=1)

        # Display selected data
        self.selected_data_display = QTextEdit(self)
        self.selected_data_display.setReadOnly(True)
        self.layout.addWidget(self.selected_data_display, stretch=1)

        # Right side layout
        self.layout_right = QVBoxLayout()

        self.graph_list = QListWidget(self)
        self.graph_list.addItems(["Line Plot", "Bar Plot", "Scatter Plot"])
        self.layout_right.addWidget(self.graph_list)

        self.x_axis_combobox = QComboBox(self)
        self.y_axis_combobox = QComboBox(self)
        self.layout_right.addWidget(self.x_axis_combobox)
        self.layout_right.addWidget(self.y_axis_combobox)

        self.plot_button = QPushButton('Plot Data', self)
        self.plot_button.clicked.connect(self.plot_data)
        self.layout_right.addWidget(self.plot_button)

        # Add right side layout to the main layout
        self.layout.addLayout(self.layout_right, stretch=1)

        self.selected_data_matrix = None
        self.selected_data_columns = []  # Store column names

        self.graph_list.itemClicked.connect(self.update_combobox)

        self.setWindowTitle('Excel Viewer')
        self.setGeometry(100, 100, 1000, 600)

    def load_excel(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx);;All Files (*)")

        if file_name:
            self.load_data_from_excel(file_name)

    def load_data_from_excel(self, file_name):
        workbook = load_workbook(file_name, data_only=True)
        sheet = workbook.active

        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_index, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.table_widget.setItem(row_index - 1, col_index, item)
        # Resize columns to fit contents
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    def read_selected_data(self):
        selected_items = self.table_widget.selectedItems()

        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select some cells before reading data.')
            return

        rows = set()
        cols = set()

        for item in selected_items:
            rows.add(item.row())
            cols.add(item.column())

        selected_rows = sorted(list(rows))
        selected_cols = sorted(list(cols))

        self.selected_data_matrix = []
        for row in selected_rows:
            row_data = [self.table_widget.item(row, col).text() for col in selected_cols]
            self.selected_data_matrix.append(row_data)

        # Store column names along with the data matrix
        self.selected_data_columns = [f'Col{col+1}' for col in selected_cols]

        # Update combo boxes with column names
        self.x_axis_combobox.clear()
        self.x_axis_combobox.addItems(self.selected_data_columns)

        self.y_axis_combobox.clear()
        self.y_axis_combobox.addItems(self.selected_data_columns)

        # Display selected data in QTextEdit
        selected_data_str = f'Selected Data in Matrix Form:\n\n{pd.DataFrame(self.selected_data_matrix, columns=self.selected_data_columns)}'
        self.selected_data_display.setPlainText(selected_data_str)

    def update_combobox(self):
        selected_graph = self.graph_list.currentItem().text()

        if selected_graph == "Scatter Plot":
            self.x_axis_combobox.setEnabled(True)
        else:
            self.x_axis_combobox.setEnabled(False)

    def plot_data(self):
        if self.selected_data_matrix is None:
            QMessageBox.warning(self, 'No Data', 'Please read and select data before plotting.')
            return

        selected_item = self.graph_list.currentItem()

        if selected_item is None:
            QMessageBox.warning(self, 'No Graph Selected', 'Please select a graph type before plotting.')
            return

        selected_graph = selected_item.text()

        x_label = self.x_axis_combobox.currentText()
        y_label = self.y_axis_combobox.currentText()

        df = pd.DataFrame(self.selected_data_matrix, columns=self.selected_data_columns)

        if selected_graph == "Line Plot":
            self.plot_line_chart(df, x_label, y_label)
        elif selected_graph == "Bar Plot":
            self.plot_bar_chart(df, x_label, y_label)
        elif selected_graph == "Scatter Plot":
            self.plot_scatter_chart(df, x_label, y_label)

    def plot_line_chart(self, df, x_label, y_label):
        plt.figure()
        plt.plot(df[x_label], df[y_label])
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.title('Line Plot')
        plt.show()

    def plot_bar_chart(self, df, x_label, y_label):
        plt.figure()
        plt.bar(df[x_label], df[y_label])
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.title('Bar Plot')
        plt.show()

    def plot_scatter_chart(self, df, x_label, y_label):
        plt.figure()
        plt.scatter(df[x_label], df[y_label])
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.title('Scatter Plot')
        plt.show()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
